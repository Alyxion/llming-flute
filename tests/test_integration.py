"""
End-to-end integration tests.

Require running Docker stack:
    docker compose up -d
    REDIS_URL=redis://localhost:6399/0 pytest tests/test_integration.py
"""

import textwrap

import pytest
from flute.client import SessionClient

pytestmark = pytest.mark.integration


@pytest.fixture
async def c():
    client = SessionClient()
    await client._ensure_connected()
    return client


# ---------------------------------------------------------------------------
# Typical ChatGPT-style apps
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_basic_execution(c):
    sid = await c.submit("print('hello world')", session_id="test-basic")
    res = await c.wait(sid)
    assert res["status"] == "completed"
    assert "hello world" in res["logs"]


@pytest.mark.asyncio
async def test_input_files(c):
    code = textwrap.dedent("""\
        with open("input.txt") as f:
            data = f.read()
        with open("output.txt", "w") as f:
            f.write(data.upper())
    """)
    sid = await c.submit(
        code, session_id="test-files", input_files={"input.txt": b"hello from input file"}
    )
    res = await c.wait(sid)
    assert res["status"] == "completed"
    assert res["output_files"]["output.txt"] == b"HELLO FROM INPUT FILE"


@pytest.mark.asyncio
async def test_matplotlib_chart(c):
    code = textwrap.dedent("""\
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
        x = np.linspace(0, 2 * np.pi, 200)
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.plot(x, np.sin(x), label="sin(x)")
        ax.plot(x, np.cos(x), label="cos(x)")
        ax.set_title("Trigonometric Functions")
        ax.legend()
        ax.grid(True)
        fig.savefig("chart.png", dpi=150)
        print("chart saved")
    """)
    sid = await c.submit(
        code, session_id="test-matplotlib", max_runtime_seconds=30, max_memory_mb=256
    )
    res = await c.wait(sid, timeout=30)
    assert res["status"] == "completed"
    png = res["output_files"]["chart.png"]
    assert png[:8] == b"\x89PNG\r\n\x1a\n"
    assert len(png) > 10_000


@pytest.mark.asyncio
async def test_pandas_csv_analysis(c):
    csv_data = (
        b"name,age,city\nAlice,30,Berlin\nBob,25,Munich\n"
        b"Charlie,35,Berlin\nDiana,28,Hamburg\nEve,32,Berlin"
    )
    code = textwrap.dedent("""\
        import pandas as pd
        df = pd.read_csv("people.csv")
        summary = df.groupby("city").agg(count=("name","count"), avg_age=("age","mean"))
        summary.reset_index().to_csv("summary.csv", index=False)
        print("done")
    """)
    sid = await c.submit(
        code, session_id="test-pandas", input_files={"people.csv": csv_data}, max_memory_mb=256
    )
    res = await c.wait(sid, timeout=30)
    assert res["status"] == "completed"
    assert "Berlin" in res["output_files"]["summary.csv"].decode()


@pytest.mark.asyncio
async def test_numpy_computation(c):
    code = textwrap.dedent("""\
        import numpy as np
        np.random.seed(42)
        A = np.random.randn(500, 500)
        B = np.random.randn(500, 500)
        C = A @ B
        print(f"C shape: {C.shape}")
        np.save("result.npy", C)
    """)
    sid = await c.submit(code, session_id="test-numpy", max_memory_mb=256)
    res = await c.wait(sid, timeout=30)
    assert res["status"] == "completed"
    assert "C shape: (500, 500)" in res["logs"]
    assert "result.npy" in res["output_files"]


@pytest.mark.asyncio
async def test_excel_generation(c):
    code = textwrap.dedent("""\
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales"
        for col, h in enumerate(["Product","Q1","Q2","Total"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
        data = [("Widget",100,200), ("Gadget",300,400)]
        for row_idx, (name, q1, q2) in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=q1)
            ws.cell(row=row_idx, column=3, value=q2)
            ws.cell(row=row_idx, column=4, value=q1+q2)
        wb.save("report.xlsx")
    """)
    sid = await c.submit(code, session_id="test-excel", max_memory_mb=256)
    res = await c.wait(sid, timeout=30)
    assert res["status"] == "completed"
    assert res["output_files"]["report.xlsx"][:2] == b"PK"


@pytest.mark.asyncio
async def test_multi_plot_report(c):
    code = textwrap.dedent("""\
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
        np.random.seed(0)
        fig, axes = plt.subplots(2, 2, figsize=(12, 10))
        axes[0,0].hist(np.random.normal(0,1,1000), bins=30)
        axes[0,0].set_title("Histogram")
        x,y = np.random.randn(2,200)
        axes[0,1].scatter(x, y, alpha=0.5)
        axes[0,1].set_title("Scatter")
        axes[1,0].bar(["A","B","C"], [10,20,30])
        axes[1,0].set_title("Bar")
        t = np.linspace(0,10,100)
        axes[1,1].plot(t, np.sin(t))
        axes[1,1].set_title("Line")
        fig.tight_layout()
        fig.savefig("report.png", dpi=150)
    """)
    sid = await c.submit(
        code, session_id="test-multiplot", max_runtime_seconds=30, max_memory_mb=256
    )
    res = await c.wait(sid, timeout=30)
    assert res["status"] == "completed"
    assert res["output_files"]["report.png"][:8] == b"\x89PNG\r\n\x1a\n"
    assert len(res["output_files"]["report.png"]) > 30_000


@pytest.mark.asyncio
async def test_image_processing(c):
    code = textwrap.dedent("""\
        from PIL import Image, ImageDraw
        img = Image.new("RGB", (400, 200), color=(30, 30, 30))
        draw = ImageDraw.Draw(img)
        draw.rectangle([20, 20, 380, 180], outline="white", width=3)
        draw.text((50, 80), "Hello from Flute!", fill="cyan")
        rotated = img.rotate(15, expand=True, fillcolor=(30, 30, 30))
        rotated.save("greeting.png")
    """)
    sid = await c.submit(code, session_id="test-pillow", max_memory_mb=256)
    res = await c.wait(sid, timeout=30)
    assert res["status"] == "completed"
    assert res["output_files"]["greeting.png"][:8] == b"\x89PNG\r\n\x1a\n"


@pytest.mark.asyncio
async def test_sympy_math(c):
    code = textwrap.dedent("""\
        from sympy import symbols, integrate, solve, sin, cos
        x = symbols("x")
        expr = sin(x)**2 + cos(x)**2
        print(f"sin^2+cos^2 = {expr.simplify()}")
        roots = solve(x**3 - 6*x**2 + 11*x - 6, x)
        print(f"roots = {roots}")
        with open("results.txt","w") as f:
            f.write(f"roots: {roots}\\n")
    """)
    sid = await c.submit(
        code, session_id="test-sympy", max_runtime_seconds=30, max_memory_mb=256
    )
    res = await c.wait(sid, timeout=30)
    assert res["status"] == "completed"
    assert "sin^2+cos^2 = 1" in res["logs"]
    assert "[1, 2, 3]" in res["logs"]
    assert "results.txt" in res["output_files"]


@pytest.mark.asyncio
async def test_multiple_output_files(c):
    code = textwrap.dedent("""\
        for i in range(3):
            with open(f"result_{i}.txt", "w") as f:
                f.write(f"result {i}")
    """)
    sid = await c.submit(code, session_id="test-multi-out")
    res = await c.wait(sid)
    assert res["status"] == "completed"
    assert len(res["output_files"]) == 3


# ---------------------------------------------------------------------------
# Error handling
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_syntax_error(c):
    sid = await c.submit("def foo(\n", session_id="test-syntax")
    res = await c.wait(sid)
    assert res["status"] == "error"
    assert "SyntaxError" in res["logs"]


@pytest.mark.asyncio
async def test_exception(c):
    sid = await c.submit("raise ValueError('test error')", session_id="test-exception")
    res = await c.wait(sid)
    assert res["status"] == "error"
    assert "test error" in res["logs"]


# ---------------------------------------------------------------------------
# Bust tests: disk
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_bust_disk_single_large_file(c):
    code = textwrap.dedent("""\
        with open("bigfile.bin", "wb") as f:
            for i in range(500):
                f.write(b"X" * (1024 * 1024))
                f.flush()
        print("SHOULD NOT REACH HERE")
    """)
    sid = await c.submit(
        code, session_id="test-bust-disk-single", max_disk_mb=10, max_runtime_seconds=30
    )
    res = await c.wait(sid, timeout=35)
    assert res["status"] in ("killed", "error")
    assert "SHOULD NOT REACH HERE" not in res["logs"]


@pytest.mark.asyncio
async def test_bust_disk_many_small_files(c):
    code = textwrap.dedent("""\
        for i in range(200):
            with open(f"chunk_{i:04d}.bin", "wb") as f:
                f.write(b"Y" * (1024 * 1024))
        print("SHOULD NOT REACH HERE")
    """)
    sid = await c.submit(
        code, session_id="test-bust-disk-many", max_disk_mb=10, max_runtime_seconds=30
    )
    res = await c.wait(sid, timeout=35)
    assert res["status"] in ("killed", "error")
    assert "SHOULD NOT REACH HERE" not in res["logs"]


@pytest.mark.asyncio
async def test_bust_disk_nested_dirs(c):
    code = textwrap.dedent("""\
        import os, time
        for depth in range(50):
            path = os.path.join(*[f"d{d}" for d in range(depth + 1)])
            os.makedirs(path, exist_ok=True)
            for j in range(3):
                with open(os.path.join(path, f"data_{j}.bin"), "wb") as f:
                    f.write(b"Z" * (1024 * 1024))
            time.sleep(0.05)
        print("SHOULD NOT REACH HERE")
    """)
    sid = await c.submit(
        code, session_id="test-bust-disk-nested", max_disk_mb=10, max_runtime_seconds=30
    )
    res = await c.wait(sid, timeout=35)
    assert res["status"] in ("killed", "error")
    assert "SHOULD NOT REACH HERE" not in res["logs"]


@pytest.mark.asyncio
async def test_bust_disk_tmpfs_preserved(c):
    bust_code = textwrap.dedent("""\
        with open("bigfile.bin", "wb") as f:
            for i in range(100):
                f.write(b"A" * (1024 * 1024))
                f.flush()
    """)
    sid1 = await c.submit(
        bust_code, session_id="test-disk-cleanup-bust", max_disk_mb=5, max_runtime_seconds=15
    )
    await c.wait(sid1, timeout=20)

    check_code = textwrap.dedent("""\
        import shutil
        usage = shutil.disk_usage("/tmp")
        used_mb = usage.used / (1024 * 1024)
        assert used_mb < 50, f"tmpfs not cleaned: {used_mb:.1f} MB"
        print("DISK CLEAN")
    """)
    sid2 = await c.submit(
        check_code, session_id="test-disk-cleanup-verify", max_memory_mb=256
    )
    res2 = await c.wait(sid2, timeout=15)
    assert res2["status"] == "completed"
    assert "DISK CLEAN" in res2["logs"]


# ---------------------------------------------------------------------------
# Bust tests: memory
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_bust_memory_fast_alloc(c):
    code = textwrap.dedent("""\
        blocks = []
        for i in range(200):
            blocks.append(bytearray(10 * 1024 * 1024))
        print("SHOULD NOT REACH HERE")
    """)
    sid = await c.submit(
        code, session_id="test-bust-mem-fast", max_memory_mb=64, max_runtime_seconds=15
    )
    res = await c.wait(sid, timeout=20)
    assert res["status"] in ("killed", "error")
    assert "SHOULD NOT REACH HERE" not in res["logs"]


@pytest.mark.asyncio
async def test_bust_memory_mmap(c):
    code = textwrap.dedent("""\
        import mmap
        maps = []
        for i in range(100):
            m = mmap.mmap(-1, 10 * 1024 * 1024)
            m.write(b"X" * (10 * 1024 * 1024))
            maps.append(m)
        print("SHOULD NOT REACH HERE")
    """)
    sid = await c.submit(
        code, session_id="test-bust-mem-mmap", max_memory_mb=64, max_runtime_seconds=15
    )
    res = await c.wait(sid, timeout=20)
    assert res["status"] in ("killed", "error")
    assert "SHOULD NOT REACH HERE" not in res["logs"]


# ---------------------------------------------------------------------------
# Bust tests: loops
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_bust_loop_busy_spin(c):
    code = "i = 0\nwhile True:\n    i += 1"
    sid = await c.submit(code, session_id="test-bust-loop-spin", max_runtime_seconds=3)
    res = await c.wait(sid, timeout=15)
    assert res["status"] == "killed"
    assert "runtime" in res["error"].lower()


@pytest.mark.asyncio
async def test_bust_loop_sleep(c):
    code = "import time\nwhile True:\n    time.sleep(0.01)"
    sid = await c.submit(code, session_id="test-bust-loop-sleep", max_runtime_seconds=3)
    res = await c.wait(sid, timeout=15)
    assert res["status"] == "killed"
    assert "runtime" in res["error"].lower()


@pytest.mark.asyncio
async def test_bust_loop_fork_bomb(c):
    code = textwrap.dedent("""\
        import subprocess, sys
        children = []
        for i in range(200):
            try:
                p = subprocess.Popen([sys.executable, "-c", "import time; time.sleep(60)"])
                children.append(p.pid)
            except Exception as e:
                print(f"fork blocked at {i}: {e}")
                break
        print(f"spawned {len(children)} children")
    """)
    sid = await c.submit(
        code, session_id="test-bust-fork", max_runtime_seconds=10, max_memory_mb=256
    )
    res = await c.wait(sid, timeout=20)
    assert res["status"] in ("killed", "error", "completed")
    if res["status"] == "completed":
        assert "fork blocked" in res["logs"]


@pytest.mark.asyncio
async def test_bust_loop_subprocess_chain(c):
    code = textwrap.dedent("""\
        import subprocess, sys
        subprocess.Popen(
            [sys.executable, "-c", "import time; time.sleep(999)"],
            start_new_session=True,
        )
        import time; time.sleep(999)
    """)
    sid = await c.submit(code, session_id="test-bust-orphan", max_runtime_seconds=3)
    res = await c.wait(sid, timeout=15)
    assert res["status"] in ("killed", "error")


# ---------------------------------------------------------------------------
# Security: env var isolation
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_env_vars_cleared(c):
    """User code cannot see REDIS_URL or other server env vars."""
    code = textwrap.dedent("""\
        import os
        redis_url = os.environ.get("REDIS_URL", "NOT FOUND")
        print(f"REDIS_URL={redis_url}")
        with open("env.txt", "w") as f:
            f.write(redis_url)
    """)
    sid = await c.submit(code, session_id="test-env-cleared")
    res = await c.wait(sid, timeout=10)
    assert res["status"] == "completed"
    assert res["output_files"]["env.txt"] == b"NOT FOUND"


# ---------------------------------------------------------------------------
# Security: cannot destroy container filesystem
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_cannot_delete_system_files(c):
    """Attempting to delete system files fails (read-only filesystem)."""
    code = textwrap.dedent("""\
        import os, shutil
        errors = []
        for path in ["/app", "/usr", "/etc"]:
            try:
                shutil.rmtree(path)
                errors.append(f"deleted {path}")
            except Exception as e:
                print(f"blocked: {path}: {e}")
        if errors:
            raise RuntimeError(f"Should not have deleted: {errors}")
        print("ALL BLOCKED")
    """)
    sid = await c.submit(
        code, session_id="test-no-delete-sys", max_runtime_seconds=10
    )
    res = await c.wait(sid, timeout=15)
    assert res["status"] == "completed"
    assert "ALL BLOCKED" in res["logs"]


@pytest.mark.asyncio
async def test_workdir_restricted(c):
    """Workdir has restricted permissions (0o700)."""
    code = textwrap.dedent("""\
        import os, stat
        cwd = os.getcwd()
        mode = stat.S_IMODE(os.stat(cwd).st_mode)
        print(f"workdir mode: {oct(mode)}")
        assert mode == 0o700, f"Expected 0o700, got {oct(mode)}"
        print("PERMISSIONS OK")
    """)
    sid = await c.submit(code, session_id="test-workdir-perms")
    res = await c.wait(sid, timeout=10)
    assert res["status"] == "completed"
    assert "PERMISSIONS OK" in res["logs"]


# ---------------------------------------------------------------------------
# Quota tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_quota_fresh_user_has_full_budget(c):
    """A new user with no prior usage has full quota available."""
    q = await c.quota("quota-fresh")
    assert q["user_id"] == "quota-fresh"
    assert q["worker_type"] == "python"
    assert q["used_seconds"] == 0.0
    assert q["remaining_seconds"] == q["limit_seconds"]
    assert q["resets_in_seconds"] == -1


@pytest.mark.asyncio
async def test_quota_tracks_usage(c):
    """Running a session charges the user's quota."""
    user = "quota-track"
    # Clear any previous quota
    await c.r.delete(f"quota:{user}:python")

    sid = await c.submit(
        "import time; time.sleep(1); print('done')",
        session_id="test-quota-track",
        user_id=user,
        max_runtime_seconds=10,
    )
    res = await c.wait(sid, timeout=15)
    assert res["status"] == "completed"

    q = await c.quota(user)
    assert q["used_seconds"] >= 1.0
    assert q["remaining_seconds"] < q["limit_seconds"]
    assert q["resets_in_seconds"] > 0


@pytest.mark.asyncio
async def test_quota_accumulates_across_sessions(c):
    """Multiple sessions accumulate usage under the same user."""
    user = "quota-accum"
    await c.r.delete(f"quota:{user}:python")

    for i in range(3):
        sid = await c.submit(
            "import time; time.sleep(0.5)",
            session_id=f"test-quota-accum-{i}",
            user_id=user,
            max_runtime_seconds=10,
        )
        await c.wait(sid, timeout=10)

    q = await c.quota(user)
    assert q["used_seconds"] >= 1.5


@pytest.mark.asyncio
async def test_quota_exceeded_rejects_session(c):
    """When quota is exhausted, the session is immediately rejected."""
    user = "quota-exceeded"
    # Artificially exhaust the quota
    rules_str = await c.r.get("quota:__config:rules") or "*:600/3600"
    from flute.quota_rules import get_quota_for_type, parse_quota_rules

    rules = parse_quota_rules(rules_str)
    limit, _ = get_quota_for_type(rules, "python")
    await c.r.set(f"quota:{user}:python", str(limit + 100))
    await c.r.expire(f"quota:{user}:python", 3600)

    sid = await c.submit(
        "print('should not run')",
        session_id="test-quota-exceeded",
        user_id=user,
    )
    res = await c.wait(sid, timeout=10)

    # --- assert well-defined rejection response ---
    assert res["status"] == "quota_exceeded"
    assert res["exit_code"] is None  # no process was spawned
    assert res["logs"] == ""  # no output
    assert res["output_files"] == {}  # no files
    assert "quota exceeded" in res["error"].lower()
    # Error message contains limit info
    assert str(limit) in res["error"]


@pytest.mark.asyncio
async def test_quota_exceeded_shows_in_quota_query(c):
    """client.quota() reflects the exceeded state."""
    user = "quota-exceeded-query"
    rules_str = await c.r.get("quota:__config:rules") or "*:600/3600"
    from flute.quota_rules import get_quota_for_type, parse_quota_rules

    rules = parse_quota_rules(rules_str)
    limit, _ = get_quota_for_type(rules, "python")
    await c.r.set(f"quota:{user}:python", str(limit + 50))
    await c.r.expire(f"quota:{user}:python", 1800)

    q = await c.quota(user)
    assert q["used_seconds"] >= limit
    assert q["remaining_seconds"] == 0.0
    assert q["resets_in_seconds"] > 0


@pytest.mark.asyncio
async def test_quota_no_user_id_bypasses_quota(c):
    """Sessions without user_id are not subject to quota."""
    sid = await c.submit(
        "print('no quota')",
        session_id="test-quota-bypass",
    )
    res = await c.wait(sid, timeout=10)
    assert res["status"] == "completed"
    assert "no quota" in res["logs"]


@pytest.mark.asyncio
async def test_quota_independent_per_user(c):
    """Different users have independent quota counters."""
    user_a, user_b = "quota-indep-a", "quota-indep-b"
    for u in (user_a, user_b):
        await c.r.delete(f"quota:{u}:python")

    # Only user_a runs a session
    sid = await c.submit(
        "import time; time.sleep(1)",
        session_id="test-quota-indep",
        user_id=user_a,
        max_runtime_seconds=10,
    )
    await c.wait(sid, timeout=15)

    qa = await c.quota(user_a)
    qb = await c.quota(user_b)
    assert qa["used_seconds"] >= 1.0
    assert qb["used_seconds"] == 0.0


@pytest.mark.asyncio
async def test_quota_charged_on_killed_session(c):
    """Quota is charged even when a session is killed for exceeding limits."""
    user = "quota-killed"
    await c.r.delete(f"quota:{user}:python")

    sid = await c.submit(
        "import time; time.sleep(60)",
        session_id="test-quota-killed",
        user_id=user,
        max_runtime_seconds=2,
    )
    res = await c.wait(sid, timeout=15)
    assert res["status"] == "killed"

    q = await c.quota(user)
    assert q["used_seconds"] >= 2.0


@pytest.mark.asyncio
async def test_quota_charged_on_error_session(c):
    """Quota is charged even when the session code raises an error."""
    user = "quota-error"
    await c.r.delete(f"quota:{user}:python")

    sid = await c.submit(
        "raise RuntimeError('boom')",
        session_id="test-quota-error",
        user_id=user,
    )
    res = await c.wait(sid, timeout=10)
    assert res["status"] == "error"

    q = await c.quota(user)
    assert q["used_seconds"] > 0


# ---------------------------------------------------------------------------
# Worker discovery
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_workers_visible(c):
    """At least one python worker should be visible via heartbeat."""
    workers = await c.workers()
    python_workers = [w for w in workers if w["worker_type"] == "python"]
    assert len(python_workers) >= 1
    assert all(w["alive"] for w in python_workers)


# ---------------------------------------------------------------------------
# Image worker tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_image_grayscale(c):
    """Image worker converts a color image to grayscale."""
    import io

    from PIL import Image

    # Create a test image
    img = Image.new("RGB", (100, 80), color=(255, 0, 0))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    img_bytes = buf.getvalue()

    sid = await c.submit(
        "",  # no code for image worker
        session_id="test-img-grayscale",
        worker_type="image",
        operation="grayscale",
        input_files={"photo.png": img_bytes},
    )
    res = await c.wait(sid, timeout=15)
    assert res["status"] == "completed"
    assert "grayscale" in res["logs"].lower()
    assert "output.png" in res["output_files"]

    # Verify output is actually grayscale
    out_img = Image.open(io.BytesIO(res["output_files"]["output.png"]))
    assert out_img.mode == "L"


@pytest.mark.asyncio
async def test_image_resize(c):
    """Image worker resizes an image."""
    import io

    from PIL import Image

    img = Image.new("RGB", (200, 150), color=(0, 128, 255))
    buf = io.BytesIO()
    img.save(buf, format="JPEG")
    img_bytes = buf.getvalue()

    sid = await c.submit(
        "",
        session_id="test-img-resize",
        worker_type="image",
        operation="resize",
        params={"width": 50, "height": 40},
        input_files={"photo.jpg": img_bytes},
    )
    res = await c.wait(sid, timeout=15)
    assert res["status"] == "completed"
    assert "50x40" in res["logs"]

    out_img = Image.open(io.BytesIO(res["output_files"]["output.jpg"]))
    assert out_img.size == (50, 40)


@pytest.mark.asyncio
async def test_image_thumbnail(c):
    """Image worker creates a thumbnail."""
    import io

    from PIL import Image

    img = Image.new("RGB", (400, 300), color=(0, 255, 0))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    img_bytes = buf.getvalue()

    sid = await c.submit(
        "",
        session_id="test-img-thumbnail",
        worker_type="image",
        operation="thumbnail",
        params={"max_size": 100},
        input_files={"photo.png": img_bytes},
    )
    res = await c.wait(sid, timeout=15)
    assert res["status"] == "completed"
    assert "100" in res["logs"]

    out_img = Image.open(io.BytesIO(res["output_files"]["output.png"]))
    assert max(out_img.size) <= 100
